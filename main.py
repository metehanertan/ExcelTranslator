import xlsxwriter as xlsxwriter
from translate import Translator
import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfilename
import os

def translate(str):
    translator = Translator(to_lang="tr")
    translation = translator.translate(str)
    return translation

def translate_excel():
    excel_path = askopenfilename()

    if excel_path == "":
        return

    name = excel_path.split("\\")[len(excel_path.split("\\")) - 1]
    new_name = name.replace(".", "_tr.")
    new_loc = excel_path.replace(name, new_name)

    workbook = openpyxl.load_workbook(excel_path)
    new_workbook = xlsxwriter.Workbook(new_loc)

    for sheet in workbook.worksheets:
        worksheet = new_workbook.add_worksheet()

        for row in range(0, sheet.max_row):
            col_num = 0
            for col in sheet.iter_cols(1, sheet.max_column):
                if not col[row].value is None:
                    worksheet.write(row, col_num, translate(col[row].value))
                col_num += 1
    workbook.close()
    new_workbook.close()
    os.startfile(new_loc)

window = tk.Tk()
window.title('Excel Translator')

greeting = tk.Label(text="Çevirisi yapılacak Excel dosyasını seçin:")
greeting.pack()

save_button = tk.Button(text="Çevir ve Kaydet", width=42, command=translate_excel)
save_button.pack()

window.mainloop()
